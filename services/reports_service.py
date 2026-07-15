"""FitLife — Reports Service + CSV Export (Phase 6)"""
import csv, io, logging
from typing import Optional
from datetime import date
from database.connection import DatabaseConnection

logger = logging.getLogger(__name__)


def get_member_report(branch_id: Optional[int] = None,
                      status: Optional[str] = None) -> list:
    try:
        sql = """
            SELECT m.full_name, m.cnic, m.phone, m.email,
                   b.branch_name, t.full_name AS trainer,
                   mp.plan_name, m.join_date, m.expiry_date,
                   m.status, m.fitness_goal, m.weight_kg, m.bmi
            FROM   members m
            LEFT JOIN branches b ON m.branch_id=b.id
            LEFT JOIN trainers t ON m.trainer_id=t.id
            LEFT JOIN membership_plans mp ON m.membership_plan_id=mp.id
            WHERE  1=1
        """
        params = []
        if branch_id: sql += " AND m.branch_id=?"; params.append(branch_id)
        if status:    sql += " AND m.status=?";    params.append(status)
        sql += " ORDER BY m.full_name"
        with DatabaseConnection() as (conn, cursor):
            cursor.execute(sql, params)
            return cursor.fetchall()
    except Exception as e:
        logger.error(f"get_member_report: {e}"); return []


def get_payment_report(branch_id: Optional[int] = None,
                       date_from: Optional[date] = None,
                       date_to:   Optional[date] = None) -> list:
    try:
        sql = """
            SELECT m.full_name, m.phone, b.branch_name,
                   py.invoice_number, py.amount, py.payment_date,
                   py.payment_method, py.status, py.notes
            FROM   payments py
            JOIN   members m ON py.member_id=m.id
            JOIN   branches b ON m.branch_id=b.id
            WHERE  1=1
        """
        params = []
        if branch_id: sql += " AND m.branch_id=?"; params.append(branch_id)
        if date_from: sql += " AND py.payment_date>=?"; params.append(date_from)
        if date_to:   sql += " AND py.payment_date<=?"; params.append(date_to)
        sql += " ORDER BY py.payment_date DESC"
        with DatabaseConnection() as (conn, cursor):
            cursor.execute(sql, params)
            return cursor.fetchall()
    except Exception as e:
        logger.error(f"get_payment_report: {e}"); return []


def get_attendance_report(branch_id: Optional[int] = None,
                          date_from: Optional[date] = None,
                          date_to:   Optional[date] = None) -> list:
    try:
        sql = """
            SELECT m.full_name, b.branch_name,
                   a.date, a.check_in_time, a.check_out_time, a.status
            FROM   attendance a
            JOIN   members m ON a.member_id=m.id
            JOIN   branches b ON m.branch_id=b.id
            WHERE  1=1
        """
        params = []
        if branch_id: sql += " AND m.branch_id=?"; params.append(branch_id)
        if date_from: sql += " AND a.date>=?"; params.append(date_from)
        if date_to:   sql += " AND a.date<=?"; params.append(date_to)
        sql += " ORDER BY a.date DESC, m.full_name"
        with DatabaseConnection() as (conn, cursor):
            cursor.execute(sql, params)
            return cursor.fetchall()
    except Exception as e:
        logger.error(f"get_attendance_report: {e}"); return []


def get_trainer_performance_report(branch_id: Optional[int] = None) -> list:
    try:
        sql = """
            SELECT t.full_name, b.branch_name, t.specialization,
                   COUNT(m.id) AS members,
                   ISNULL(t.performance_rating, 0) AS rating,
                   t.monthly_salary, t.hire_date, t.status
            FROM   trainers t
            JOIN   branches b ON t.branch_id=b.id
            LEFT JOIN members m ON m.trainer_id=t.id AND m.status='Active'
            WHERE  1=1
        """
        params = []
        if branch_id: sql += " AND t.branch_id=?"; params.append(branch_id)
        sql += " GROUP BY t.id, t.full_name, b.branch_name, t.specialization, t.performance_rating, t.monthly_salary, t.hire_date, t.status ORDER BY members DESC"
        with DatabaseConnection() as (conn, cursor):
            cursor.execute(sql, params)
            return cursor.fetchall()
    except Exception as e:
        logger.error(f"get_trainer_performance_report: {e}"); return []


def get_equipment_report(branch_id: Optional[int] = None) -> list:
    try:
        sql = """
            SELECT e.equipment_name, e.category, e.brand, e.model,
                   e.serial_number, e.condition_status, e.quantity,
                   e.purchase_price, e.purchase_date,
                   e.last_maintenance_date, b.branch_name
            FROM   equipment e
            JOIN   branches b ON e.branch_id=b.id
            WHERE  1=1
        """
        params = []
        if branch_id: sql += " AND e.branch_id=?"; params.append(branch_id)
        sql += " ORDER BY e.category, e.equipment_name"
        with DatabaseConnection() as (conn, cursor):
            cursor.execute(sql, params)
            return cursor.fetchall()
    except Exception as e:
        logger.error(f"get_equipment_report: {e}"); return []


def export_to_csv(headers: list, rows: list, file_path: str) -> dict:
    try:
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows([[str(v) if v is not None else "" for v in row] for row in rows])
        return {"success": True, "message": f"Exported {len(rows)} rows to CSV successfully."}
    except Exception as e:
        logger.error(f"export_to_csv: {e}")
        return {"success": False, "message": str(e)}


def export_to_excel(title: str, headers: list, rows: list, file_path: str) -> dict:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.views.sheetView[0].showGridLines = True
        
        title_font = Font(name="Segoe UI", size=16, bold=True, color="0066FF")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        
        header_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
        zebra_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        
        thin_side = Side(border_style="thin", color="D1D5DB")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Title banner
        ws.append([title])
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 30
        ws.append([]) # Spacer row
        
        # Header Row
        ws.append(headers)
        header_row_idx = 3
        ws.row_dimensions[header_row_idx].height = 24
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            
        # Data Rows
        for row_idx, row in enumerate(rows, start=4):
            ws.append([v if v is not None else "" for v in row])
            ws.row_dimensions[row_idx].height = 20
            is_zebra = (row_idx % 2 == 0)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if is_zebra:
                    cell.fill = zebra_fill
                    
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row > 1 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(file_path)
        return {"success": True, "message": f"Exported {len(rows)} rows to Excel successfully."}
    except Exception as e:
        logger.error(f"export_to_excel: {e}")
        return {"success": False, "message": str(e)}


def export_to_pdf(title: str, headers: list, rows: list, file_path: str) -> dict:
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        # Determine page layout orientation based on column density
        page_size = landscape(letter) if len(headers) > 6 else letter
        doc = SimpleDocTemplate(file_path, pagesize=page_size,
                                rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            textColor=colors.HexColor("#0066FF"),
            spaceAfter=10
        )
        
        meta_style = ParagraphStyle(
            name="MetaStyle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=15
        )
        
        cell_header_style = ParagraphStyle(
            name="CellHeaderStyle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            textColor=colors.white
        )
        
        cell_data_style = ParagraphStyle(
            name="CellDataStyle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            textColor=colors.HexColor("#1F2937")
        )
        
        story = []
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"Generated: {date.today().strftime('%B %d, %Y')} — FitLife Management System", meta_style))
        story.append(Spacer(1, 8))
        
        # Setup table wrapper lists
        table_data = []
        table_data.append([Paragraph(h, cell_header_style) for h in headers])
        for r in rows:
            table_data.append([Paragraph(str(v) if v is not None else "", cell_data_style) for v in r])
            
        page_w = page_size[0] - 72
        col_w = page_w / len(headers)
        
        t = Table(table_data, colWidths=[col_w] * len(headers))
        t_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0066FF")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ])
        t.setStyle(t_style)
        story.append(t)
        
        doc.build(story)
        return {"success": True, "message": f"Exported {len(rows)} rows to PDF successfully."}
    except Exception as e:
        logger.error(f"export_to_pdf: {e}")
        return {"success": False, "message": str(e)}
